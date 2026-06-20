"""Extract Phase 0 external blockers from shared planning spreadsheets."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def rows_by_header(path: Path, header_first_cell: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, keep_vba=path.suffix == ".xlsm")
    worksheet = workbook.worksheets[0]
    headers = None
    records: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        if headers is None:
            if values and values[0] == header_first_cell:
                headers = [str(value) if value is not None else "" for value in values]
            continue
        if not any(value is not None for value in values):
            continue
        records.append(
            {
                header: values[index] if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
        )
    return records


def is_unresolved(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).upper()
    return "MANAGEMENT INPUT REQUIRED" in text or text in {"TBC", "TODO"}


def extract_external_blockers(drive_root: str | Path) -> dict[str, Any]:
    """Read spreadsheets and report decisions that code cannot resolve."""

    drive_root = Path(drive_root)
    shared_root = drive_root / "00_Shared"
    validation_path = drive_root / "01_Group_ReID_Role_Tracking" / "Validation_Results.xlsx"
    blockers: list[dict[str, Any]] = []
    warnings: list[str] = []

    data_catalogue_path = shared_root / "Data_Catalogue.xlsx"
    if data_catalogue_path.exists():
        for row in rows_by_header(data_catalogue_path, "Dataset ID"):
            dataset_id = row.get("Dataset ID")
            if dataset_id in {"DS-001", "DS-002"}:
                unresolved_fields = [
                    field
                    for field in ["Ground Truth Available", "Owner", "Drive Link / Path", "Capture Date"]
                    if is_unresolved(row.get(field)) or row.get(field) in {None, "No"}
                ]
                if unresolved_fields:
                    blockers.append(
                        {
                            "area": "dataset_access",
                            "source": str(data_catalogue_path),
                            "item": dataset_id,
                            "unresolved_fields": unresolved_fields,
                            "row": row,
                        }
                    )
    else:
        blockers.append(
            {
                "area": "dataset_access",
                "source": str(data_catalogue_path),
                "item": "Data_Catalogue.xlsx missing",
            }
        )

    questions_path = shared_root / "Open_Questions_and_TODOs.xlsm"
    if questions_path.exists():
        for row in rows_by_header(questions_path, "Area"):
            area = row.get("Area")
            if area in {"Dataset access", "Ground truth availability"} and is_unresolved(row.get("Status")):
                blockers.append(
                    {
                        "area": str(area).lower().replace(" ", "_"),
                        "source": str(questions_path),
                        "item": row.get("Question / Decision Needed"),
                        "status": row.get("Status"),
                        "owner": row.get("Owner"),
                        "due": row.get("Due"),
                    }
                )
    else:
        blockers.append(
            {
                "area": "open_questions",
                "source": str(questions_path),
                "item": "Open_Questions_and_TODOs.xlsm missing",
            }
        )

    if validation_path.exists():
        for row in rows_by_header(validation_path, "Metric"):
            metric = row.get("Metric")
            target = row.get("Target / Threshold")
            current = row.get("Current Result")
            if metric in {
                "Cross-camera association accuracy",
                "ID switches per delivery",
                "Role classification accuracy",
            } and (is_unresolved(target) or is_unresolved(current)):
                blockers.append(
                    {
                        "area": "validation_targets",
                        "source": str(validation_path),
                        "item": metric,
                        "current_result": current,
                        "target_or_threshold": target,
                    }
                )
    else:
        blockers.append(
            {
                "area": "validation_targets",
                "source": str(validation_path),
                "item": "Validation_Results.xlsx missing",
            }
        )

    return {
        "external_readiness": "blocked" if blockers else "ready",
        "blockers": blockers,
        "recommended_decisions": [
            "Confirm DS-001 owner and local/Drive path.",
            "Select DS-002 blind validation subset.",
            "Assign one owner for manual ID and role ground-truth labels.",
            "Choose annotation tooling before treating P1 model scores as final.",
            "Set thresholds for association accuracy, ID switches, and role accuracy.",
        ],
        "required_ground_truth_labels": [
            "bbox",
            "coco_17_keypoints_2d",
            "global_player_id",
            "role",
            "optional_3d_reference",
        ],
        "warnings": warnings,
    }

